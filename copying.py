import openpyxl
import os
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange


def fill_data_into_newfile(template_path, output_path, title, cell_range):
    if not os.path.exists(template_path):
        print(f"Template file {template_path} does not exist.")
        return

    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)

    # Load workbooks
    output_wb = openpyxl.load_workbook(output_path, keep_vba=True)
    print(f"Loaded {output_path}")

    template_wb = openpyxl.load_workbook(template_path, keep_vba=True)
    template_ws = template_wb[title]
    print(f"Loaded template '{title}' from {template_path}")

    # Create sheet
    new_sheet = output_wb.create_sheet(title=title)
    print(f"Created new sheet '{title}'")


    new_sheet = output_wb.create_sheet(title=title)  # Remove existing sheet with the same title if it exists
    for row in template_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    default_column_width = 8.43  # Default width for Excel columns
    if template_ws.title == "281":
        default_column_width = 3.0
    # Copy column widths
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        dim = template_ws.column_dimensions.get(col_letter)
        if dim and dim.width is not None:
            new_sheet.column_dimensions[col_letter].width = dim.width
        else:
            new_sheet.column_dimensions[col_letter].width = default_column_width
            print(f"⚠️ Column {col_letter} has no explicit width, skipping.")

    # Copy row heights
    for row_idx in range(min_row, max_row + 1):
        dim = template_ws.row_dimensions.get(row_idx)
        if dim and dim.height is not None:
            new_sheet.row_dimensions[row_idx].height = dim.height
        else:
            print(f"⚠️ Row {row_idx} has no explicit height, skipping.")

    # Copy merged cells
    for merged_range in template_ws.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))
    
    output_wb.remove(output_wb[title])
    new_sheet.title = title  # Rename the new sheet to the desired title
    output_wb.save(output_path)
    print(f"Copied template sheet '{title}' to {output_path}")
    return