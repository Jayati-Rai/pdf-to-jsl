import openpyxl
import os
from copy import copy

def fill_data_into_template(template_path, output_path):
    if not os.path.exists(template_path):
        print(f"Template file {template_path} does not exist.")
        return

    # Load the output workbook
    output_wb = openpyxl.load_workbook(output_path, keep_vba=True)
    print(f"Loaded {output_path}")

    # Create a new sheet
    new_sheet = output_wb.create_sheet(title="प्रपत्र-ग")
    print(f"Created new sheet 'प्रपत्र-ग'")

    # Load the template workbook and sheet
    template_wb = openpyxl.load_workbook(template_path, keep_vba=True)
    template_ws = template_wb["प्रपत्र-ग"]
    print(f"Loaded template {template_path}")

    # Copy column widths
    for col_letter, col_dim in template_ws.column_dimensions.items():
        new_sheet.column_dimensions[col_letter].width = col_dim.width

    # Copy merged cells
    merged_top_left_cells = set()
    merged_all_cells = set()
    for merged_range in template_ws.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))
        merged_top_left_cells.add(merged_range.coord.split(":")[0])
        for row in template_ws[merged_range.coord]:
            for cell in row:
                merged_all_cells.add(cell.coordinate)

    # Copy cells
    for row in template_ws.iter_rows():
        for cell in row:
            coord = cell.coordinate
            new_cell = new_sheet[coord]

            try:
                # Copy only if not in a merged range or if it's the top-left of a merged range
                if coord not in merged_all_cells or coord in merged_top_left_cells:
                    new_cell.value = cell.value  # Includes formulas
                    print(f"Copied cell {coord} with value: {cell.value}")
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
            except Exception as e:
                print(f"Failed to copy cell {coord}: {e}")

    # Save final workbook
    output_wb.save(output_path)
    print(f"प्रपत्र-ग copied and filled into {output_path}")